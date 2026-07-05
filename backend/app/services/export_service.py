from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.trade import Trade


def trades_to_workbook(sheets: dict[str, list[Trade]], summaries: dict[str, dict]) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    for title, trades in sheets.items():
        ws = wb.create_sheet(title[:31])
        write_trade_sheet(ws, trades)

    if summaries:
        ws = wb.create_sheet("汇总统计")
        headers = ["名称", "本金", "账户净值", "累计盈亏", "收益率", "胜率", "交易次数", "总盈利", "总亏损", "最大回撤", "累计R", "期望R"]
        headers[-1] = "平均 R"
        headers.extend(["平均风险比例", "超出2%笔数", "计划执行率", "未按计划笔数", "2%模拟净值", "2%模拟收益率"])
        ws.append(headers)
        style_header(ws)
        for name, item in summaries.items():
            ws.append([
                name,
                item["initial_capital"],
                item["account_equity"],
                item["total_pnl"],
                item["total_return"],
                item["win_rate"],
                item["total_trades"],
                item["total_profit"],
                item["total_loss"],
                item["max_drawdown"],
                item["total_r"],
                item["average_r"],
                item["average_risk_percent"],
                item["over_risk_count"],
                item["plan_adherence_rate"],
                item["deviation_count"],
                item["simulated_equity"],
                item["simulated_return"],
            ])
        autosize(ws)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def write_trade_sheet(ws, trades: list[Trade]) -> None:
    headers = [
        "平仓时间", "交易对", "方向", "开仓价", "平仓价", "止损价", "止盈价",
        "净盈亏", "风险比例", "计划风险金额", "R倍数", "计划盈亏比",
        "账户收益贡献", "交易前净值", "开仓时间", "持仓秒数", "策略标签",
        "离场原因", "状态", "备注",
    ]
    headers[9] = "计划止损金额"
    headers.extend([
        "计划止盈金额", "是否按计划执行", "偏离原因分类", "偏离原因说明",
        "2%模拟风险金额", "2%模拟止盈金额", "2%模拟实际盈亏",
    ])
    ws.append(headers)
    style_header(ws)
    for trade in trades:
        ws.append([
            trade.close_time,
            trade.symbol,
            trade.direction.value,
            trade.avg_open_price,
            trade.avg_close_price,
            trade.stop_loss_price,
            trade.take_profit_price,
            trade.realized_pnl,
            trade.risk_percent,
            trade.planned_risk_amount,
            trade.r_multiple,
            trade.planned_rr,
            trade.account_return_percent,
            trade.account_equity_before,
            trade.open_time,
            trade.holding_seconds,
            trade.strategy_tag or "",
            trade.exit_reason or "",
            trade.status.value,
            trade.note or "",
            trade.planned_profit_amount,
            "是" if trade.followed_plan else "否",
            trade.deviation_reason or "",
            trade.deviation_note or "",
            trade.simulated_risk_amount,
            trade.simulated_profit_target,
            trade.simulated_pnl,
        ])
    autosize(ws)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def autosize(ws) -> None:
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(width, 28)
